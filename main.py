# main.py
from fastapi import FastAPI, HTTPException, Request, Response, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict
import sqlite3
import openpyxl
from datetime import datetime
import os
import uuid
import re
import logging
import zipfile
import io

app = FastAPI(title="Streamer Nominations API")

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Настройка CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# Модели Pydantic
class Vote(BaseModel):
    username: str
    telegram_id: Optional[int] = None
    nomination: str
    nominee: str
    is_custom: bool = False
    request_id: str = None

class CustomVote(BaseModel):
    username: str
    telegram_id: Optional[int] = None
    nomination: str
    custom_nominee: str
    request_id: str = None

class SearchRequest(BaseModel):
    query: str
    nomination: str

class UserRegistration(BaseModel):
    telegram_id: int
    username: str
    first_name: str
    last_name: str

class AdminCreate(BaseModel):
    username: str
    password: str

# Глобальный словарь для хранения номинантов по номинациям
nominees_by_nomination: Dict[str, List[str]] = {}

# Инициализация БД
def init_db():
    conn = sqlite3.connect('votes.db')
    c = conn.cursor()
    
    # Таблица для голосов
    c.execute('''CREATE TABLE IF NOT EXISTS votes
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT NOT NULL,
                  telegram_id INTEGER NOT NULL,
                  nomination TEXT NOT NULL,
                  nominee TEXT NOT NULL,
                  is_custom BOOLEAN DEFAULT FALSE,
                  request_id TEXT UNIQUE,
                  timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                  UNIQUE(telegram_id, nomination))''')  # Уникальный индекс
    
    # Таблица для пользователей
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  telegram_id INTEGER UNIQUE,
                  username TEXT,
                  first_name TEXT,
                  last_name TEXT,
                  created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
    
    # Таблица для админов
    c.execute('''CREATE TABLE IF NOT EXISTS admins
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE NOT NULL,
                  telegram_id INTEGER UNIQUE,
                  created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
    
    # Таблица для логов действий админов
    c.execute('''CREATE TABLE IF NOT EXISTS admin_logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  admin_id INTEGER NOT NULL,
                  action TEXT NOT NULL,
                  details TEXT,
                  timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                  FOREIGN KEY(admin_id) REFERENCES admins(id))''')
    
    # Индексы для улучшения производительности
    c.execute('''CREATE INDEX IF NOT EXISTS idx_votes_request_id ON votes(request_id)''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_votes_user_nomination ON votes(telegram_id, nomination)''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_users_telegram_id ON users(telegram_id)''')
    c.execute('''CREATE INDEX IF NOT EXISTS idx_admins_telegram_id ON admins(telegram_id)''')
    
    conn.commit()
    conn.close()

    # Создаем Excel файлы если не существуют
    if not os.path.exists('votes.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Голоса"
        ws.append(['ID пользователя', 'Ник пользователя', 'Номинация', 'Номинант', 'Время голосования'])
        wb.save('votes.xlsx')
    
    if not os.path.exists('users.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Пользователи"
        ws.append(['Telegram ID', 'Username', 'Имя', 'Фамилия', 'Дата регистрации'])
        wb.save('users.xlsx')
    
    # Загрузка допустимых номинантов из файла (если есть)
    load_allowed_nominees()

def load_allowed_nominees():
    """Загружает допустимых номинантов из текстового файла с учетом номинаций"""
    global nominees_by_nomination
    
    try:
        if os.path.exists('allowed_nominees.txt'):
            with open('allowed_nominees.txt', 'r', encoding='utf-8') as f:
                content = f.read()
                
                # Очищаем словарь
                nominees_by_nomination = {}
                
                # Разделяем содержимое по двойному переводу строки (между номинациями)
                nominations_blocks = re.split(r'\n\s*\n', content.strip())
                
                for block in nominations_blocks:
                    lines = block.strip().split('\n')
                    if not lines:
                        continue
                    
                    # Первая строка - название номинации (должна заканчиваться на :)
                    nomination_line = lines[0].strip()
                    if not nomination_line.endswith(':'):
                        continue
                    
                    nomination = nomination_line[:-1].strip()  # Убираем двоеточие
                    nominees = []
                    
                    # Остальные строки - номинанты
                    for line in lines[1:]:
                        line = line.strip()
                        # Пропускаем пустые строки и комментарии
                        if line and not line.startswith('#') and not line.startswith('//'):
                            nominees.append(line)
                    
                    if nominees:  # Добавляем только если есть номинанты
                        nominees_by_nomination[nomination] = nominees
                        logger.info(f"Загружено {len(nominees)} номинантов для номинации '{nomination}'")
                    
    except Exception as e:
        logger.error(f"Ошибка при загрузке номинантов: {e}")

# Функция проверки прав администратора
async def verify_admin(telegram_id: int = Header(None, alias="X-Telegram-ID")):
    if not telegram_id:
        raise HTTPException(status_code=401, detail="Требуется идентификатор Telegram")
    
    conn = sqlite3.connect('votes.db')
    c = conn.cursor()
    c.execute("SELECT id FROM admins WHERE telegram_id = ?", (telegram_id,))
    admin = c.fetchone()
    conn.close()
    
    if not admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    
    return True

# Функция логирования действий админов
# Функция логирования действий админов
def log_admin_action(admin_id: int, action: str, details: str = ""):
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        c.execute("INSERT INTO admin_logs (admin_id, action, details) VALUES (?, ?, ?)",
                 (admin_id, action, details))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.error(f"Ошибка при логировании действия администратора: {e}")

# Инициализация при старте
init_db()

# Роуты
@app.post("/register-user")
async def register_user(user: UserRegistration, request: Request):
    try:
        logger.info(f"Получен запрос на регистрацию: {user.model_dump()}")
        
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        
        # Проверяем, существует ли уже пользователь
        c.execute("SELECT id, username, first_name, last_name FROM users WHERE telegram_id = ?", (user.telegram_id,))
        existing_user = c.fetchone()
        
        if existing_user:
            # Проверяем, изменились ли данные пользователя
            existing_username, existing_first_name, existing_last_name = existing_user[1], existing_user[2], existing_user[3]
            
            if (existing_username != user.username or 
                existing_first_name != user.first_name or 
                existing_last_name != user.last_name):
                
                # Обновляем данные пользователя, если они изменились
                c.execute("UPDATE users SET username = ?, first_name = ?, last_name = ? WHERE telegram_id = ?",
                         (user.username, user.first_name, user.last_name, user.telegram_id))
                message = "Данные пользователя обновлены"
                needs_excel_update = True
            else:
                # Данные не изменились, просто возвращаем успех
                message = "Пользователь уже зарегистрирован"
                needs_excel_update = False
        else:
            # Создаем нового пользователя
            c.execute("INSERT INTO users (telegram_id, username, first_name, last_name) VALUES (?, ?, ?, ?)",
                     (user.telegram_id, user.username, user.first_name, user.last_name))
            message = "Пользователь зарегистрирован"
            needs_excel_update = True
        
        conn.commit()
        conn.close()
        
        # Обновляем Excel только если нужно
        if needs_excel_update:
            try:
                wb = openpyxl.load_workbook('users.xlsx')
                ws = wb.active
                
                # Проверяем, существует ли уже пользователь в Excel
                user_exists = False
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value == user.telegram_id:
                        # Обновляем данные существующего пользователя
                        ws.cell(row=row, column=2, value=user.username)
                        ws.cell(row=row, column=3, value=user.first_name)
                        ws.cell(row=row, column=4, value=user.last_name)
                        ws.cell(row=row, column=5, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                        user_exists = True
                        break
                
                if not user_exists:
                    # Добавляем нового пользователя
                    ws.append([user.telegram_id, user.username, user.first_name, user.last_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                
                wb.save('users.xlsx')
            except Exception as e:
                logger.error(f"Ошибка при сохранении в Excel: {e}")
        
        return {"message": message, "user": user.model_dump()}
    except Exception as e:
        logger.error(f"Ошибка при регистрации пользователя: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/user-votes/{telegram_id}")
async def get_user_votes(telegram_id: int):
    """Получает все голоса пользователя по его telegram_id"""
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        c.execute("SELECT nomination, nominee FROM votes WHERE telegram_id = ?", (telegram_id,))
        votes = c.fetchall()
        conn.close()
        
        return {vote[0]: vote[1] for vote in votes}
    except Exception as e:
        logger.error(f"Ошибка при получении голосов пользователя: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/vote")
async def add_vote(vote: Vote):
    conn = None  # Инициализируем переменную заранее
    try:
        # Фильтруем фантомные запросы с дефолтными значениями
        if (vote.username == "Default User" or 
            vote.telegram_id is None or 
            str(vote.telegram_id) == "N/A"):
            logger.warning(f"Отклонен фантомный запрос: {vote.model_dump()}")
            return {"message": "Фантомный запрос отклонен"}
        
        if not vote.request_id:
            vote.request_id = str(uuid.uuid4())
        
        nominee_display = f"СВОЙ ВАРИАНТ({vote.nominee})" if vote.is_custom else vote.nominee
        
        conn = sqlite3.connect('votes.db')
        conn.execute("PRAGMA busy_timeout = 5000")
        c = conn.cursor()
        
        # Начинаем транзакцию
        conn.execute("BEGIN IMMEDIATE")
        
        # Проверяем существующий голос
        c.execute("""
            SELECT id FROM votes 
            WHERE nomination = ? AND (telegram_id = ? OR username = ?)
        """, (vote.nomination, vote.telegram_id, vote.username))
        
        existing_vote = c.fetchone()
        
        if existing_vote:
            conn.rollback()
            conn.close()
            logger.info(f"Обнаружен существующий голос: nomination={vote.nomination}, user={vote.telegram_id or vote.username}")
            return {"message": "Вы уже голосовали в этой номинации"}
        
        # Проверяем, не был ли уже обработан этот запрос
        c.execute("SELECT id FROM votes WHERE request_id = ?", (vote.request_id,))
        existing_request = c.fetchone()
        
        if existing_request:
            conn.rollback()
            conn.close()
            return {"message": "Голос уже учтен"}
        
        # Добавляем новый голос
        c.execute("""
            INSERT INTO votes (username, telegram_id, nomination, nominee, is_custom, request_id) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (vote.username, vote.telegram_id, vote.nomination, nominee_display, vote.is_custom, vote.request_id))
        
        conn.commit()
        conn.close()

        # Добавляем в Excel
        try:
            wb = openpyxl.load_workbook('votes.xlsx')
            ws = wb.active
            ws.append([vote.telegram_id, vote.username, vote.nomination, nominee_display, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            wb.save('votes.xlsx')
        except Exception as e:
            logger.error(f"Ошибка при сохранении в Excel: {e}")

        logger.info(f"Успешный голос: {vote.username} в номинации {vote.nomination}")
        return {"message": "Голос успешно сохранен"}
        
    except sqlite3.IntegrityError as e:
        if "UNIQUE constraint" in str(e):
            return {"message": "Голос уже учтен"}
        logger.error(f"Ошибка целостности базы данных: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Ошибка при обработке голоса: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/revote")
async def revote(vote: Vote):
    """Переголосование: удаляет старый голос и добавляет новый"""
    conn = None
    try:
        # Фильтруем фантомные запросы
        if (vote.username == "Default User" or 
            vote.telegram_id is None or 
            str(vote.telegram_id) == "N/A"):
            logger.warning(f"Отклонен фантомный запрос: {vote.model_dump()}")
            return {"message": "Фантомный запрос отклонен"}
        
        if not vote.request_id:
            vote.request_id = str(uuid.uuid4())
        
        nominee_display = f"СВОЙ ВАРИАНТ({vote.nominee})" if vote.is_custom else vote.nominee
        
        conn = sqlite3.connect('votes.db')
        conn.execute("PRAGMA busy_timeout = 5000")
        c = conn.cursor()
        
        # Начинаем транзакцию
        conn.execute("BEGIN IMMEDIATE")
        
        # Удаляем старый голос в этой номинации
        c.execute("DELETE FROM votes WHERE nomination = ? AND telegram_id = ?", 
                 (vote.nomination, vote.telegram_id))
        
        # Добавляем новый голос
        c.execute("""
            INSERT INTO votes (username, telegram_id, nomination, nominee, is_custom, request_id) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (vote.username, vote.telegram_id, vote.nomination, nominee_display, vote.is_custom, vote.request_id))
        
        conn.commit()
        conn.close()

        # Обновляем Excel: удаляем старую запись и добавляем новую
        try:
            wb = openpyxl.load_workbook('votes.xlsx')
            ws = wb.active
            
            # Ищем и удаляем старую запись
            for row in range(2, ws.max_row + 1):
                if (ws.cell(row=row, column=1).value == vote.telegram_id and 
                    ws.cell(row=row, column=3).value == vote.nomination):
                    ws.delete_rows(row)
                    break
            
            # Добавляем новую запись
            ws.append([vote.telegram_id, vote.username, vote.nomination, nominee_display, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            wb.save('votes.xlsx')
        except Exception as e:
            logger.error(f"Ошибка при обновлении Excel: {e}")

        logger.info(f"Успешное переголосование: {vote.username} в номинации {vote.nomination}")
        return {"message": "Переголосование успешно выполнено"}
        
    except sqlite3.IntegrityError as e:
        if "UNIQUE constraint" in str(e):
            return {"message": "Голос уже учтен"}
        logger.error(f"Ошибка целостности базы данных: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Ошибка при обработке переголосования: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/vote-custom")
async def add_custom_vote(vote: CustomVote):
    conn = None  # Инициализируем переменную заранее
    try:
        # Фильтруем фантомные запросы с дефолтными значениями
        if (vote.username == "Default User" or 
            vote.telegram_id is None or 
            str(vote.telegram_id) == "N/A"):
            logger.warning(f"Отклонен фантомный запрос: {vote.model_dump()}")
            return {"message": "Фантомный запрос отклонен"}
        
        if not vote.request_id:
            vote.request_id = str(uuid.uuid4())
        
        if not vote.custom_nominee.strip():
            raise HTTPException(status_code=400, detail="Имя номинанта не может быть пустым")
        
        # Проверяем, что выбранный номинант допустим для этой номинации
        if vote.nomination in nominees_by_nomination:
            allowed_nominees = nominees_by_nomination[vote.nomination]
            if vote.custom_nominee not in allowed_nominees:
                raise HTTPException(status_code=400, detail="Этот номинант недопустим для данной номинации")
        
        nominee_display = f"СВОЙ ВАРИАНТ({vote.custom_nominee})"
        
        conn = sqlite3.connect('votes.db')
        conn.execute("PRAGMA busy_timeout = 5000")
        c = conn.cursor()
        
        # Начинаем транзакцию
        conn.execute("BEGIN IMMEDIATE")
        
        # Проверяем существующий голос
        c.execute("""
            SELECT id FROM votes 
            WHERE nomination = ? AND (telegram_id = ? OR username = ?)
        """, (vote.nomination, vote.telegram_id, vote.username))
        
        existing_vote = c.fetchone()
        
        if existing_vote:
            conn.rollback()
            conn.close()
            logger.info(f"Обнаружен существующий голос: nomination={vote.nomination}, user={vote.telegram_id or vote.username}")
            return {"message": "Вы уже голосовали в этой номинации"}
        
        # Проверяем, не был ли уже обработан этот запрос
        c.execute("SELECT id FROM votes WHERE request_id = ?", (vote.request_id,))
        existing_request = c.fetchone()
        
        if existing_request:
            conn.rollback()
            conn.close()
            return {"message": "Голос уже учтен"}
        
        # Добавляем новый голос
        c.execute("""
            INSERT INTO votes (username, telegram_id, nomination, nominee, is_custom, request_id) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (vote.username, vote.telegram_id, vote.nomination, nominee_display, True, vote.request_id))
        
        conn.commit()
        conn.close()

        # Добавляем в Excel
        try:
            wb = openpyxl.load_workbook('votes.xlsx')
            ws = wb.active
            ws.append([vote.telegram_id, vote.username, vote.nomination, nominee_display, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            wb.save('votes.xlsx')
        except Exception as e:
            logger.error(f"Ошибка при сохранении в Excel: {e}")

        logger.info(f"Успешный кастомный голос: {vote.username} в номинации {vote.nomination}")
        return {"message": "Голос за свой вариант успешно сохранен"}
        
    except sqlite3.IntegrityError as e:
        if "UNIQUE constraint" in str(e):
            return {"message": "Голос уже учтен"}
        logger.error(f"Ошибка целостности базы данных: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Ошибка при обработке кастомного голоса: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/revote-custom")
async def revote_custom(vote: CustomVote):
    """Переголосование для кастомного номинанта"""
    conn = None
    try:
        # Фильтруем фантомные запросы
        if (vote.username == "Default User" or 
            vote.telegram_id is None or 
            str(vote.telegram_id) == "N/A"):
            logger.warning(f"Отклонен фантомный запрос: {vote.model_dump()}")
            return {"message": "Фантомный запрос отклонен"}
        
        if not vote.request_id:
            vote.request_id = str(uuid.uuid4())
        
        if not vote.custom_nominee.strip():
            raise HTTPException(status_code=400, detail="Имя номинанта не может быть пустым")
        
        # Проверяем, что выбранный номинант допустим
        if vote.nomination in nominees_by_nomination:
            allowed_nominees = nominees_by_nomination[vote.nomination]
            if vote.custom_nominee not in allowed_nominees:
                raise HTTPException(status_code=400, detail="Этот номинант недопустим для данной номинации")
        
        nominee_display = f"СВОЙ ВАРИАНТ({vote.custom_nominee})"
        
        conn = sqlite3.connect('votes.db')
        conn.execute("PRAGMA busy_timeout = 5000")
        c = conn.cursor()
        
        # Начинаем транзакцию
        conn.execute("BEGIN IMMEDIATE")
        
        # Удаляем старый голос в этой номинации
        c.execute("DELETE FROM votes WHERE nomination = ? AND telegram_id = ?", 
                 (vote.nomination, vote.telegram_id))
        
        # Добавляем новый голос
        c.execute("""
            INSERT INTO votes (username, telegram_id, nomination, nominee, is_custom, request_id) 
            VALUES (?, ?, ?, ?, ?, ?)
        """, (vote.username, vote.telegram_id, vote.nomination, nominee_display, True, vote.request_id))
        
        conn.commit()
        conn.close()

        # Обновляем Excel
        try:
            wb = openpyxl.load_workbook('votes.xlsx')
            ws = wb.active
            
            # Ищем и удаляем старую запись
            for row in range(2, ws.max_row + 1):
                if (ws.cell(row=row, column=1).value == vote.telegram_id and 
                    ws.cell(row=row, column=3).value == vote.nomination):
                    ws.delete_rows(row)
                    break
            
            # Добавляем новую запись
            ws.append([vote.telegram_id, vote.username, vote.nomination, nominee_display, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            wb.save('votes.xlsx')
        except Exception as e:
            logger.error(f"Ошибка при обновлении Excel: {e}")

        logger.info(f"Успешное кастомное переголосование: {vote.username} в номинации {vote.nomination}")
        return {"message": "Переголосование за свой вариант успешно выполнено"}
        
    except sqlite3.IntegrityError as e:
        if "UNIQUE constraint" in str(e):
            return {"message": "Голос уже учтен"}
        logger.error(f"Ошибка целостности базы данных: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Ошибка при обработке кастомного переголосования: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/search-nominees")
async def search_nominees(request: SearchRequest):
    """Поиск номинантов по запросу и номинации"""
    try:
        if not request.query or len(request.query) < 2:
            return {"results": []}
        
        results = []
        
        # Ищем совпадения в допустимых номинантах для конкретной номинации
        if request.nomination in nominees_by_nomination:
            allowed_nominees = nominees_by_nomination[request.nomination]
            query_lower = request.query.lower()
            
            for nominee in allowed_nominees:
                if query_lower in nominee.lower():
                    results.append(nominee)
                    if len(results) >= 10:  # Ограничиваем количество результатов
                        break
        
        return {"results": results}
    except Exception as e:
        logger.error(f"Ошибка при поиске номинантов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/results")
async def get_results():
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        c.execute("SELECT nomination, nominee, COUNT(*) as votes FROM votes GROUP BY nomination, nominee")
        results = c.fetchall()
        conn.close()
        
        return [{"nomination": r[0], "nominee": r[1], "votes": r[2]} for r in results]
    except Exception as e:
        logger.error(f"Ошибка при получении результатов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/votes")
async def get_all_votes():
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        c.execute("SELECT * FROM votes ORDER BY timestamp DESC")
        votes = c.fetchall()
        conn.close()
        
        return votes
    except Exception as e:
        logger.error(f"Ошибка при получении всех голосов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/clean-invalid-votes")
async def clean_invalid_votes():
    """Очищает некорректные голоса из базы данных"""
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        
        # Удаляем голоса с некорректными данными
        c.execute('''
            DELETE FROM votes 
            WHERE telegram_id IS NULL 
            OR username = 'Default User'
            OR telegram_id = 'N/A'
        ''')
        
        deleted_count = c.rowcount
        conn.commit()
        conn.close()
        
        return {"message": f"Удалено {deleted_count} некорректных записей"}
    except Exception as e:
        logger.error(f"Ошибка при очистке некорректных голосов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/clean-phantom-votes")
async def clean_phantom_votes():
    """Очищает фантомные голоса из базы данных"""
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        
        # Удаляем фантомные голоса
        c.execute('''
            DELETE FROM votes 
            WHERE username = 'Default User' 
            OR telegram_id IS NULL 
            OR telegram_id = 'N/A'
        ''')
        
        deleted_count = c.rowcount
        conn.commit()
        conn.close()
        
        return {"message": f"Удалено {deleted_count} фантомных записей"}
    except Exception as e:
        logger.error(f"Ошибка при очистке фантомных голосов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/users")
async def get_all_users():
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        c.execute("SELECT * FROM users ORDER BY created_at DESC")
        users = c.fetchall()
        conn.close()
        
        return users
    except Exception as e:
        logger.error(f"Ошибка при получении всех пользователей: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/reload-nominees")
async def reload_nominees():
    """Перезагружает список допустимых номинантов из файла"""
    try:
        load_allowed_nominees()
        return {"message": "Список номинантов успешно обновлен"}
    except Exception as e:
        logger.error(f"Ошибка при перезагрузке номинантов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/add")
async def add_admin(admin: AdminCreate):
    try:
        # Проверяем пароль
        if admin.password != "jarvis2023":
            raise HTTPException(status_code=401, detail="Неверный пароль")
        
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        
        # Получаем telegram_id из users по username
        c.execute("SELECT telegram_id FROM users WHERE username = ?", (admin.username,))
        user = c.fetchone()
        
        if not user:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        
        telegram_id = user[0]
        
        # Проверяем, не является ли уже админом
        c.execute("SELECT id FROM admins WHERE username = ? OR telegram_id = ?", 
                 (admin.username, telegram_id))
        existing_admin = c.fetchone()
        
        if existing_admin:
            raise HTTPException(status_code=400, detail="Пользователь уже является администратором")
        
        # Добавляем админа
        c.execute("INSERT INTO admins (username, telegram_id) VALUES (?, ?)", 
                 (admin.username, telegram_id))
        
        # Логируем действие (без указания кто добавил, так как нет информации об инициаторе)
        log_admin_action(
            0,  # ID 0 для действий по паролю
            "ADD_ADMIN_BY_PASSWORD", 
            f"Добавлен администратор: {admin.username}"
        )
        
        conn.commit()
        conn.close()
        
        logger.info(f"Добавлен новый администратор по паролю: {admin.username}")
        return {"message": f"Пользователь {admin.username} добавлен как администратор"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при добавлении администратора: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/download-data", dependencies=[Depends(verify_admin)])
async def download_data():
    try:
        # Создаем архив в памяти
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Добавляем файлы votes.xlsx и users.xlsx
            for filename in ['votes.xlsx', 'users.xlsx']:
                if os.path.exists(filename):
                    with open(filename, 'rb') as f:
                        zip_file.writestr(filename, f.read())
            
            # Создаем дамп базы данных SQLite
            conn = sqlite3.connect('votes.db')
            dump_data = '\n'.join(conn.iterdump())
            zip_file.writestr('database_dump.sql', dump_data)
            conn.close()
        
        zip_buffer.seek(0)
        
        # Формируем имя файла с датой и временем
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"voting_data_{current_time}.zip"
        
        # Возвращаем архив как ответ
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/zip"
            }
        )
        
    except Exception as e:
        logger.error(f"Ошибка при создании архива: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/check")
async def check_admin(telegram_id: int):
    try:
        conn = sqlite3.connect('votes.db')
        c = conn.cursor()
        c.execute("SELECT username FROM admins WHERE telegram_id = ?", (telegram_id,))
        admin = c.fetchone()
        conn.close()
        
        return {"is_admin": bool(admin)}
        
    except Exception as e:
        logger.error(f"Ошибка при проверке прав администратора: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)